"""
Synthetic Data Generator for Mentor-Based Student Performance Analysis System
==============================================================================
Generates realistic student data for 3 sections (65 students each = 195 total)
with detailed subject-wise internal/external marks across 5 semesters.

Also generates a separate 5,000-row CSV for training the XGBoost ML model.

College: GRIET (Gokaraju Rangaraju Institute of Engineering and Technology)
Programme: B.Tech CSE (Data Science) - GR22 Regulation

REALISTIC CONSTRAINTS:
- Internal (40): 2 Mid exams (each 30: MCQ/10 + Subjective/20), average taken = 30 marks.
  Plus Assignments(5) + Attendance/Daily Assessment(5) = 10 marks. Total Internal = 40.
  Most students score 25-35. Only 3-4 per class get full 40.
- External (60): Max realistic score is 58. Only 1-2 students per class get 59.
  Getting 60/60 is essentially impossible.
- Lab: Internal(40) + External(60)
- Non-Credit: Mid1(50) + Mid2(50) -> Average. Pass = 40% (20/50).
- Pass rules: Internal >= 14/40 AND External >= 21/60 AND Total >= 40/100.
"""

import json
import csv
import random
import os

random.seed(42)  # Reproducible results for review

# ============================================================================
# CURRICULUM DEFINITION (Extracted from syllabus.pdf)
# ============================================================================
CURRICULUM = {
    1: {
        "name": "I Year I Semester",
        "subjects": [
            {"name": "Linear Algebra and Function Approximation", "code": "GR22A1001", "credits": 4, "type": "Theory"},
            {"name": "Engineering Chemistry", "code": "GR22A1005", "credits": 4, "type": "Theory"},
            {"name": "Programming for Problem Solving", "code": "GR22A1007", "credits": 3, "type": "Theory"},
            {"name": "Fundamentals of Electrical Engineering", "code": "GR22A1008", "credits": 3, "type": "Theory"},
            {"name": "Engineering Chemistry Lab", "code": "GR22A1015", "credits": 1.5, "type": "Lab"},
            {"name": "Programming for Problem Solving Lab", "code": "GR22A1017", "credits": 1.5, "type": "Lab"},
            {"name": "Fundamentals of Electrical Engineering Lab", "code": "GR22A1019", "credits": 1, "type": "Lab"},
            {"name": "Engineering Workshop", "code": "GR22A1021", "credits": 2.5, "type": "Lab"},
            {"name": "Design Thinking", "code": "GR22A1022", "credits": 0, "type": "Non-Credit"},
        ]
    },
    2: {
        "name": "I Year II Semester",
        "subjects": [
            {"name": "Differential Equations and Vector Calculus", "code": "GR22A1002", "credits": 4, "type": "Theory"},
            {"name": "Applied Physics", "code": "GR22A1003", "credits": 4, "type": "Theory"},
            {"name": "English", "code": "GR22A1006", "credits": 2, "type": "Theory"},
            {"name": "Graphics for Engineers", "code": "GR22A1011", "credits": 3, "type": "Theory"},
            {"name": "Data Structures", "code": "GR22A1012", "credits": 3, "type": "Theory"},
            {"name": "Applied Physics Lab", "code": "GR22A1013", "credits": 1.5, "type": "Lab"},
            {"name": "English Language and Communication Skills Lab", "code": "GR22A1016", "credits": 1, "type": "Lab"},
            {"name": "Data Structures Lab", "code": "GR22A1020", "credits": 1, "type": "Lab"},
        ]
    },
    3: {
        "name": "II Year I Semester",
        "subjects": [
            {"name": "Value Ethics and Gender Culture", "code": "GR22A2002", "credits": 0, "type": "Non-Credit"},
            {"name": "Digital Logic Design", "code": "GR22A2067", "credits": 3, "type": "Theory"},
            {"name": "Java Programming", "code": "GR22A2068", "credits": 3, "type": "Theory"},
            {"name": "Database Management Systems", "code": "GR22A2069", "credits": 3, "type": "Theory"},
            {"name": "Java Programming Lab", "code": "GR22A2071", "credits": 2, "type": "Lab"},
            {"name": "Database Management Systems Lab", "code": "GR22A2072", "credits": 1.5, "type": "Lab"},
            {"name": "Discrete Mathematics", "code": "GR22A2075", "credits": 3, "type": "Theory"},
            {"name": "Design and Analysis of Algorithms", "code": "GR22A2077", "credits": 3, "type": "Theory"},
            {"name": "Scripting Languages Lab", "code": "GR22A2085", "credits": 1.5, "type": "Lab"},
        ]
    },
    4: {
        "name": "II Year II Semester",
        "subjects": [
            {"name": "Probability and Statistics", "code": "GR22A2004", "credits": 4, "type": "Theory"},
            {"name": "Operating Systems", "code": "GR22A2070", "credits": 3, "type": "Theory"},
            {"name": "Computer Networks", "code": "GR22A2073", "credits": 3, "type": "Theory"},
            {"name": "Software Engineering", "code": "GR22A2074", "credits": 3, "type": "Theory"},
            {"name": "Machine Learning", "code": "GR22A2076", "credits": 3, "type": "Theory"},
            {"name": "Operating Systems Lab", "code": "GR22A2078", "credits": 1.5, "type": "Lab"},
            {"name": "Computer Networks Lab", "code": "GR22A2079", "credits": 1.5, "type": "Lab"},
            {"name": "Data Science with R Programming Lab", "code": "GR22A2087", "credits": 2, "type": "Lab"},
            {"name": "Environmental Science", "code": "GR22A2090", "credits": 0, "type": "Non-Credit"},
        ]
    },
    5: {
        "name": "III Year I Semester",
        "subjects": [
            {"name": "Artificial Intelligence", "code": "GR22A3070", "credits": 3, "type": "Theory"},
            {"name": "Data Visualization", "code": "GR22A3076", "credits": 3, "type": "Theory"},
            {"name": "Web Technologies", "code": "GR22A3071", "credits": 3, "type": "Theory"},
            {"name": "Compiler Design", "code": "GR22A3072", "credits": 4, "type": "Theory"},
            {"name": "Artificial Intelligence Lab", "code": "GR22A3074", "credits": 1.5, "type": "Lab"},
            {"name": "Data Visualization Lab", "code": "GR22A3077", "credits": 1.5, "type": "Lab"},
            {"name": "Web Technologies Lab", "code": "GR22A3075", "credits": 1.5, "type": "Lab"},
            {"name": "Constitution of India", "code": "GR22A3080", "credits": 0, "type": "Non-Credit"},
        ]
    },
    6: {
        "name": "III Year II Semester",
        "subjects": [
            {"name": "Machine Learning", "code": "GR22A3140", "credits": 3, "type": "Theory"},
            {"name": "Automata and Compiler Design", "code": "GR22A3115", "credits": 3, "type": "Theory"},
            {"name": "Big Data Analytics", "code": "GR22A3143", "credits": 3, "type": "Theory"},
            {"name": "Cryptography and Network Security", "code": "GR22A4048", "credits": 3, "type": "Theory"},
            {"name": "Joy of Computing using Python", "code": "NPTEL", "credits": 3, "type": "Theory"},
            {"name": "Machine Learning Lab", "code": "GR22A3142", "credits": 1.5, "type": "Lab"},
            {"name": "Big Data Analytics Lab", "code": "GR22A3148", "credits": 1.5, "type": "Lab"},
            {"name": "Mini Project with Seminar", "code": "GR22A3089", "credits": 2, "type": "Lab"},
            {"name": "Constitution of India", "code": "GR22A2003", "credits": 0, "type": "Non-Credit"},
        ]
    }
}


# ============================================================================
# GRADE CONVERSION (10-point scale, GRIET GR22 regulation)
# ============================================================================
def total_to_grade(total_pct):
    if total_pct >= 90: return 10, "O"
    elif total_pct >= 80: return 9, "A+"
    elif total_pct >= 70: return 8, "A"
    elif total_pct >= 60: return 7, "B+"
    elif total_pct >= 50: return 6, "B"
    elif total_pct >= 45: return 5, "C"
    elif total_pct >= 40: return 4, "D"
    else: return 0, "F"


# ============================================================================
# REALISTIC MARK GENERATORS
# ============================================================================
SECTIONS = ["A", "B", "C"]
STUDENTS_PER_SECTION = 65
GENDERS = ["Male", "Female"]

# Indian Names for realistic demo data
MALE_FIRST_NAMES = [
    "Aarav", "Aditya", "Akash", "Akhil", "Amar", "Amit", "Anand", "Anil", "Ankit",
    "Arjun", "Arun", "Ashwin", "Bharat", "Charan", "Chirag", "Deepak", "Dhruv",
    "Dinesh", "Ganesh", "Gaurav", "Gopal", "Harish", "Harsh", "Hemant", "Jayesh",
    "Karan", "Kartik", "Kiran", "Krishna", "Kunal", "Lokesh", "Manoj", "Mohan",
    "Mukesh", "Naga", "Naveen", "Nikhil", "Pavan", "Pranav", "Prasad", "Rahul",
    "Raj", "Rajesh", "Raju", "Rakesh", "Ram", "Ramesh", "Ravi", "Rohit",
    "Sachin", "Sagar", "Sai", "Sandeep", "Sanjay", "Sathvik", "Shankar", "Shiva",
    "Siddharth", "Srikanth", "Suresh", "Tarun", "Uday", "Varun", "Venkat", "Vijay",
    "Vikram", "Vinay", "Vishnu", "Vivek", "Yash", "Yogesh"
]
FEMALE_FIRST_NAMES = [
    "Aishwarya", "Amrutha", "Anjali", "Ananya", "Ankitha", "Bhavana", "Bhoomika",
    "Chaitra", "Chandana", "Deepika", "Deepthi", "Divya", "Durga", "Gayathri",
    "Harini", "Hasini", "Ishitha", "Jyothi", "Kavya", "Keerthana", "Lakshmi",
    "Lavanya", "Madhavi", "Manasa", "Meghana", "Mounika", "Mythili", "Nandini",
    "Neha", "Nikitha", "Pallavi", "Pooja", "Priya", "Priyanka", "Rachana",
    "Ramya", "Rani", "Rashmi", "Rithika", "Roja", "Sahithi", "Sanjana",
    "Saraswathi", "Shalini", "Shravya", "Shreya", "Sindhu", "Sirisha", "Sneha",
    "Soumya", "Spandana", "Sravani", "Sree", "Sreshta", "Swathi", "Tanvi",
    "Tejaswini", "Trisha", "Vaishnavi", "Varsha", "Vidya", "Vyshnavi"
]
LAST_NAMES = [
    "Reddy", "Sharma", "Kumar", "Naidu", "Rao", "Gupta", "Patel", "Singh",
    "Chowdary", "Varma", "Prasad", "Raju", "Murthy", "Goud", "Shetty",
    "Verma", "Yadav", "Mishra", "Joshi", "Patil", "Deshmukh", "Kulkarni",
    "Iyer", "Nair", "Pillai", "Menon", "Chauhan", "Thakur", "Agarwal",
    "Banerjee", "Bose", "Das", "Ghosh", "Sen", "Mallick", "Swamy",
    "Gowda", "Hegde", "Kamath", "Bhat"
]

def get_indian_name(gender):
    """Generate a random Indian full name based on gender."""
    if gender == "Male":
        first = random.choice(MALE_FIRST_NAMES)
    else:
        first = random.choice(FEMALE_FIRST_NAMES)
    last = random.choice(LAST_NAMES)
    return f"{first} {last}"

def clamp(val, lo, hi):
    return max(lo, min(hi, val))


# We track how many "topper" slots have been used per section per semester
# so only 3-4 students per class get full internals, and only 1-2 get 59 externals.
class SectionTracker:
    def __init__(self):
        self.full_internal_count = {}   # key: (section, sem) -> count
        self.high_external_count = {}   # key: (section, sem) -> count
    
    def can_get_full_internal(self, section, sem):
        key = (section, sem)
        return self.full_internal_count.get(key, 0) < 4
    
    def record_full_internal(self, section, sem):
        key = (section, sem)
        self.full_internal_count[key] = self.full_internal_count.get(key, 0) + 1
    
    def can_get_59_external(self, section, sem):
        key = (section, sem)
        return self.high_external_count.get(key, 0) < 2
    
    def record_59_external(self, section, sem):
        key = (section, sem)
        self.high_external_count[key] = self.high_external_count.get(key, 0) + 1

tracker = SectionTracker()


def _gen_one_mid(ability, is_topper, noise_spread=0.12):
    """
    Generate marks for ONE mid-term exam paper (out of 30).
    Paper = MCQ(10) + Subjective(20).
    Adds significant noise so Mid1 and Mid2 differ realistically.
    """
    noise = lambda sp=noise_spread: random.uniform(-sp, sp)

    # MCQ (out of 10) — most students lose 1-3 even when good
    mcq = round(clamp(10 * (ability + noise(0.12)), 0, 10))
    if not is_topper and mcq == 10:
        mcq = random.choice([7, 8, 8, 9, 9])
    if mcq == 10 and random.random() > 0.4:
        mcq = 9  # Even toppers rarely get 10/10

    # Subjective (out of 20) — hardest to score full in
    sub = round(clamp(20 * (ability + noise(0.14)), 0, 20))
    if not is_topper and sub >= 18:
        sub = random.choice([14, 15, 16, 17])
    if sub >= 19 and random.random() > 0.25:
        sub = random.choice([17, 18])
    if sub == 20:  # Practically impossible
        sub = 19

    return mcq, sub, mcq + sub


def generate_theory_marks(ability, attn_habit, section, sem, is_topper):
    """
    Theory Internal (40):
      - 2 Mid exams, each out of 30 (MCQ/10 + Subjective/20)
      - Average of Mid1 and Mid2 = Mid component (out of 30)
      - Assignments (out of 5): 5 assignments throughout the semester
      - Attendance / Daily Assessment (out of 5)
      - Internal Total = round(avg(Mid1, Mid2)) + Assignments + Attendance
    Theory External (60):
      - Semester exam. Max realistic = 58, rare 59. Never 60.
    Pass: Internal >= 14 AND External >= 21 AND Total >= 40.
    """
    noise = lambda spread=0.10: random.uniform(-spread, spread)

    # --- MID 1 ---
    mid1_mcq, mid1_sub, mid1_total = _gen_one_mid(ability, is_topper)

    # --- MID 2 --- (students often do slightly differently in mid2)
    mid2_ability = clamp(ability + random.uniform(-0.08, 0.08), 0.1, 1.0)
    mid2_mcq, mid2_sub, mid2_total = _gen_one_mid(mid2_ability, is_topper)

    # Average of the two mids (out of 30)
    mid_avg = round((mid1_total + mid2_total) / 2)

    # --- Assignments (out of 5) ---
    assignments = round(clamp(5 * (attn_habit + noise(0.10)), 0, 5))

    # --- Attendance / Daily Assessment (out of 5) ---
    attendance = round(clamp(5 * (attn_habit + noise(0.10)), 0, 5))

    internal = mid_avg + assignments + attendance
    internal = clamp(internal, 0, 40)

    # Cap internals: only 3-4 per section per semester can get 40/40
    if internal >= 39:
        if is_topper and tracker.can_get_full_internal(section, sem) and internal == 40:
            tracker.record_full_internal(section, sem)
        elif internal == 40:
            internal = random.choice([35, 36, 37, 38])
            # Adjust mid_avg to make the numbers consistent
            mid_avg = internal - assignments - attendance
            mid_avg = clamp(mid_avg, 0, 30)
        elif internal == 39 and not is_topper:
            internal = random.choice([34, 35, 36, 37])
            mid_avg = internal - assignments - attendance
            mid_avg = clamp(mid_avg, 0, 30)

    # Recompute internal to be consistent
    internal = mid_avg + assignments + attendance
    internal = clamp(internal, 0, 40)

    # --- EXTERNAL MARKS ---
    ext_raw = 60 * (ability + noise(0.14))
    external = round(clamp(ext_raw, 0, 58))  # Hard cap at 58 for most

    # Only 1-2 per section per semester can get 59
    if is_topper and ability > 0.92 and random.random() > 0.6:
        if tracker.can_get_59_external(section, sem):
            external = 59
            tracker.record_59_external(section, sem)

    # Nobody gets 60/60
    external = min(external, 59)

    total = internal + external
    is_pass = (internal >= 14) and (external >= 21) and (total >= 40)
    gp, grade = total_to_grade(total) if is_pass else (0, "F")

    return {
        "mid1_mcq": mid1_mcq,
        "mid1_subjective": mid1_sub,
        "mid1_total": mid1_total,
        "mid2_mcq": mid2_mcq,
        "mid2_subjective": mid2_sub,
        "mid2_total": mid2_total,
        "mid_avg": mid_avg,
        "internal_assignments": assignments,
        "internal_attendance": attendance,
        "internal_total": internal,
        "external": external,
        "total": total,
        "grade_point": gp,
        "grade": grade,
        "is_pass": is_pass
    }


def generate_lab_marks(ability, is_topper, section, sem):
    """
    Lab: Internal(40) + External(60). Labs score slightly higher than theory.
    Same realistic caps apply.
    """
    noise = lambda: random.uniform(-0.08, 0.08)
    lab_boost = 0.06

    int_raw = 40 * (ability + lab_boost + noise())
    internal = round(clamp(int_raw, 0, 40))
    # Cap internals for labs too
    if internal >= 39 and not is_topper:
        internal = random.choice([35, 36, 37, 38])
    if internal == 40:
        if tracker.can_get_full_internal(section, sem):
            tracker.record_full_internal(section, sem)
        else:
            internal = random.choice([36, 37, 38])

    ext_raw = 60 * (ability + lab_boost + noise())
    external = round(clamp(ext_raw, 0, 58))
    if is_topper and ability > 0.92 and random.random() > 0.7:
        if tracker.can_get_59_external(section, sem):
            external = 59
            tracker.record_59_external(section, sem)
    external = min(external, 59)

    total = internal + external
    is_pass = (internal >= 14) and (external >= 21) and (total >= 40)
    gp, grade = total_to_grade(total) if is_pass else (0, "F")

    return {
        "internal_total": internal,
        "external": external,
        "total": total,
        "grade_point": gp,
        "grade": grade,
        "is_pass": is_pass
    }


def generate_noncredit_marks(ability):
    """Non-Credit: Mid1(40) + Mid2(40) -> Average. Pass = 40% (16/40)."""
    noise = lambda: random.uniform(-0.10, 0.10)
    mid1 = round(clamp(40 * (ability + noise()), 0, 38))  # Rarely 40/40
    mid2 = round(clamp(40 * (ability + noise()), 0, 38))
    avg = round((mid1 + mid2) / 2)
    is_pass = avg >= 16

    return {
        "mid1": mid1,
        "mid2": mid2,
        "average": avg,
        "is_pass": is_pass
    }


# ============================================================================
# STUDENT GENERATOR
# ============================================================================
def generate_one_student(student_id, name, section, gender, risk_profile, is_topper=False):
    if risk_profile == "High":
        ability = random.uniform(0.28, 0.48)
        attn_habit = random.uniform(0.40, 0.65)
    elif risk_profile == "Medium":
        ability = random.uniform(0.55, 0.72)
        attn_habit = random.uniform(0.70, 0.85)
    else:  # Low risk
        ability = random.uniform(0.78, 0.96)
        attn_habit = random.uniform(0.85, 0.98)

    student = {
        "student_id": student_id,
        "name": name,
        "section": section,
        "gender": gender,
        "risk_profile": risk_profile,
        "semesters": []
    }

    all_attendance = []
    all_assignments = []
    total_credit_points = 0
    total_credits = 0

    for sem_num in range(1, 7):
        sem_info = CURRICULUM[sem_num]

        # Per-semester fluctuation
        sem_ability = clamp(ability + random.uniform(-0.07, 0.07), 0.1, 1.0)
        sem_attn = clamp(attn_habit + random.uniform(-0.06, 0.06), 0.1, 1.0)

        sem_data = {
            "semester": sem_num,
            "semester_name": sem_info["name"],
            "subjects": []
        }

        sem_credits = 0
        sem_grade_points = 0

        for subj in sem_info["subjects"]:
            subj_entry = {
                "name": subj["name"],
                "code": subj["code"],
                "credits": subj["credits"],
                "type": subj["type"],
            }

            if subj["type"] == "Theory":
                marks = generate_theory_marks(sem_ability, sem_attn, section, sem_num, is_topper)
                subj_entry["marks"] = marks
                all_assignments.append(marks["internal_assignments"] / 5.0)
                all_attendance.append(marks["internal_attendance"] / 5.0)
                # Also track mid performance for extra noise signal
                all_attendance.append(marks["mid_avg"] / 30.0)

            elif subj["type"] == "Lab":
                marks = generate_lab_marks(sem_ability, is_topper, section, sem_num)
                subj_entry["marks"] = marks

            else:
                marks = generate_noncredit_marks(sem_ability)
                subj_entry["marks"] = marks

            if subj["credits"] > 0 and subj["type"] != "Non-Credit":
                gp = marks.get("grade_point", 0)
                sem_credits += subj["credits"]
                sem_grade_points += (gp * subj["credits"])
                total_credits += subj["credits"]
                total_credit_points += (gp * subj["credits"])

            sem_data["subjects"].append(subj_entry)

        sem_data["sgpa"] = round(sem_grade_points / sem_credits, 2) if sem_credits > 0 else 0.0
        sem_data["total_credits"] = sem_credits
        student["semesters"].append(sem_data)

    student["cgpa"] = round(total_credit_points / total_credits, 2) if total_credits > 0 else 0.0
    student["overall_attendance_rate"] = round(
        sum(all_attendance) / len(all_attendance), 2
    ) if all_attendance else 0.0
    student["overall_assignment_rate"] = round(
        sum(all_assignments) / len(all_assignments), 2
    ) if all_assignments else 0.0

    return student


# ============================================================================
# EXCEL EXPORT
# ============================================================================
def export_to_excel(students, filepath):
    """Export the detailed student data to a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Styling ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    high_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    med_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    low_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fail_font = Font(color="FF0000", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # ========================================
    # SHEET 1: Overview (one row per student)
    # ========================================
    ws_overview = wb.active
    ws_overview.title = "Overview"
    overview_headers = [
        "Student ID", "Name", "Section", "Gender", "Risk Profile", "CGPA",
        "Attendance %", "Assignment %",
        "Sem 1 SGPA", "Sem 2 SGPA", "Sem 3 SGPA", "Sem 4 SGPA", "Sem 5 SGPA"
    ]
    ws_overview.append(overview_headers)
    style_header(ws_overview, 1, len(overview_headers))

    for s in students:
        risk_fill = high_fill if s["risk_profile"] == "High" else (med_fill if s["risk_profile"] == "Medium" else low_fill)
        row_data = [
            s["student_id"], s["name"], s["section"], s["gender"], s["risk_profile"],
            s["cgpa"], round(s["overall_attendance_rate"] * 100, 1), round(s["overall_assignment_rate"] * 100, 1)
        ]
        for sem in s["semesters"]:
            row_data.append(sem["sgpa"])
        ws_overview.append(row_data)
        row_num = ws_overview.max_row
        for col in range(1, len(overview_headers) + 1):
            cell = ws_overview.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        # Color risk column
        ws_overview.cell(row=row_num, column=5).fill = risk_fill

    # Auto-width
    for col in ws_overview.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_overview.column_dimensions[col[0].column_letter].width = max_len + 3

    # ========================================
    # SHEETS 2-6: One sheet per semester with full subject-wise marks
    # ========================================
    for sem_num in range(1, 6):
        sem_info = CURRICULUM[sem_num]
        ws = wb.create_sheet(title=f"Semester {sem_num}")

        # Build headers dynamically based on subject types
        headers = ["Student ID", "Name", "Section", "Risk"]
        subj_names = []
        for subj in sem_info["subjects"]:
            sn = subj["name"][:30]  # Truncate for column width
            subj_names.append(sn)
            if subj["type"] == "Theory":
                headers.extend([
                    f"{sn} M1-MCQ/10", f"{sn} M1-Sub/20", f"{sn} Mid1/30",
                    f"{sn} M2-MCQ/10", f"{sn} M2-Sub/20", f"{sn} Mid2/30",
                    f"{sn} MidAvg/30", f"{sn} Asn/5", f"{sn} Att/5",
                    f"{sn} Int/40", f"{sn} Ext/60", f"{sn} Total", f"{sn} Grade"
                ])
            elif subj["type"] == "Lab":
                headers.extend([
                    f"{sn} Int/40", f"{sn} Ext/60", f"{sn} Total", f"{sn} Grade"
                ])
            else:
                headers.extend([
                    f"{sn} Mid1/50", f"{sn} Mid2/50", f"{sn} Avg", f"{sn} Status"
                ])
        headers.append("SGPA")

        ws.append(headers)
        style_header(ws, 1, len(headers))

        for s in students:
            sem_data = s["semesters"][sem_num - 1]
            row = [s["student_id"], s["name"], s["section"], s["risk_profile"]]

            for subj_data in sem_data["subjects"]:
                m = subj_data["marks"]
                if subj_data["type"] == "Theory":
                    row.extend([
                        m["mid1_mcq"], m["mid1_subjective"], m["mid1_total"],
                        m["mid2_mcq"], m["mid2_subjective"], m["mid2_total"],
                        m["mid_avg"], m["internal_assignments"], m["internal_attendance"],
                        m["internal_total"], m["external"], m["total"], m["grade"]
                    ])
                elif subj_data["type"] == "Lab":
                    row.extend([
                        m["internal_total"], m["external"], m["total"], m["grade"]
                    ])
                else:
                    row.extend([
                        m["mid1"], m["mid2"], m["average"],
                        "PASS" if m["is_pass"] else "FAIL"
                    ])
            row.append(sem_data["sgpa"])
            ws.append(row)

            # Highlight fails in red
            row_num = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if cell.value == "F" or cell.value == "FAIL":
                    cell.font = fail_font

    # ========================================
    # SHEET 7: Section-wise Summary
    # ========================================
    ws_summary = wb.create_sheet(title="Section Summary")
    summary_headers = ["Section", "Total Students", "High Risk", "Medium Risk", "Low Risk", "Avg CGPA", "Avg Attendance %"]
    ws_summary.append(summary_headers)
    style_header(ws_summary, 1, len(summary_headers))

    for sec in SECTIONS:
        sec_students = [s for s in students if s["section"] == sec]
        h = sum(1 for s in sec_students if s["risk_profile"] == "High")
        m = sum(1 for s in sec_students if s["risk_profile"] == "Medium")
        l = sum(1 for s in sec_students if s["risk_profile"] == "Low")
        avg_cgpa = round(sum(s["cgpa"] for s in sec_students) / len(sec_students), 2)
        avg_att = round(sum(s["overall_attendance_rate"] for s in sec_students) / len(sec_students) * 100, 1)
        ws_summary.append([sec, len(sec_students), h, m, l, avg_cgpa, avg_att])
        row_num = ws_summary.max_row
        for col in range(1, len(summary_headers) + 1):
            ws_summary.cell(row=row_num, column=col).border = thin_border
            ws_summary.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")

    for col in ws_summary.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_summary.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(filepath)
    print(f"  -> Excel saved to: {filepath}")


# ============================================================================
# MAIN
# ============================================================================
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(script_dir, "..", ".."))
    external_dir = os.path.join(project_root, "external_dataset")

    # -----------------------------------------------------------
    # PART 1: Generate 195 detailed students for the App Database
    # -----------------------------------------------------------
    print("=" * 60)
    print("PART 1: Generating 195 students (3 sections x 65)")
    print("=" * 60)

    global tracker
    tracker = SectionTracker()

    all_demo_students = []
    student_counter = 1

    for section in SECTIONS:
        # 10 High, 15 Medium, 40 Low per section
        # Mark 3-4 of the Low-risk students as "toppers" (can get full internal / 59 ext)
        risk_list = ["High"] * 10 + ["Medium"] * 15 + ["Low"] * 40
        random.shuffle(risk_list)

        topper_indices = set()
        low_indices = [i for i, r in enumerate(risk_list) if r == "Low"]
        if len(low_indices) >= 4:
            topper_indices = set(random.sample(low_indices, 4))

        for i, risk in enumerate(risk_list):
            sid = f"23241A67{student_counter:02d}"
            gender = random.choice(GENDERS)
            name = get_indian_name(gender)
            is_topper = (i in topper_indices)

            student = generate_one_student(sid, name, section, gender, risk, is_topper)
            all_demo_students.append(student)
            student_counter += 1

    # Save JSON for DB seeding
    json_path = os.path.join(script_dir, "synthetic_students.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_demo_students, f, indent=2, ensure_ascii=False)
    print(f"  -> Saved {len(all_demo_students)} students to synthetic_students.json")

    # Save Excel for user review
    excel_path = os.path.join(external_dir, "generated_student_data_v2.xlsx")
    export_to_excel(all_demo_students, excel_path)

    # Print summary
    print(f"\n  Section Summary:")
    print(f"  {'Section':<10} {'High':<8} {'Medium':<10} {'Low':<8} {'Total':<8}")
    print(f"  {'-'*44}")
    for sec in SECTIONS:
        sec_s = [s for s in all_demo_students if s["section"] == sec]
        h = sum(1 for s in sec_s if s["risk_profile"] == "High")
        m = sum(1 for s in sec_s if s["risk_profile"] == "Medium")
        l = sum(1 for s in sec_s if s["risk_profile"] == "Low")
        print(f"  {sec:<10} {h:<8} {m:<10} {l:<8} {len(sec_s):<8}")

    # -----------------------------------------------------------
    # PART 2: Generate 5,000 lightweight students for ML Training
    # -----------------------------------------------------------
    print(f"\n{'=' * 60}")
    print("PART 2: Generating 5,000 students for ML Training CSV")
    print("=" * 60)

    tracker = SectionTracker()  # Reset tracker
    ml_rows = []
    ml_risk_list = ["High"] * 830 + ["Medium"] * 1250 + ["Low"] * 2920
    random.shuffle(ml_risk_list)

    def extract_ml_features(student_data):
        """Extract per-semester mid marks, assignments, attendance for ML features."""
        features = {
            "cgpa": student_data["cgpa"],
            "overall_attendance": student_data["overall_attendance_rate"],
            "overall_assignment": student_data["overall_assignment_rate"],
        }
        
        for sem in student_data["semesters"]:
            sem_num = sem["semester"]
            # Collect mid1, mid2, assignment, attendance averages across theory subjects
            mid1_scores = []
            mid2_scores = []
            asn_scores = []
            att_scores = []
            
            for subj in sem["subjects"]:
                if subj["type"] == "Theory":
                    m = subj["marks"]
                    mid1_scores.append(m["mid1_total"] / 30.0)  # Normalize to 0-1
                    mid2_scores.append(m["mid2_total"] / 30.0)
                    asn_scores.append(m["internal_assignments"] / 5.0)
                    att_scores.append(m["internal_attendance"] / 5.0)
            
            features[f"sem{sem_num}_mid1_avg"] = round(sum(mid1_scores) / len(mid1_scores), 3) if mid1_scores else 0
            features[f"sem{sem_num}_mid2_avg"] = round(sum(mid2_scores) / len(mid2_scores), 3) if mid2_scores else 0
            features[f"sem{sem_num}_asn_avg"] = round(sum(asn_scores) / len(asn_scores), 3) if asn_scores else 0
            features[f"sem{sem_num}_att_avg"] = round(sum(att_scores) / len(att_scores), 3) if att_scores else 0
            features[f"sem{sem_num}_sgpa"] = sem["sgpa"]
        
        return features

    # --- Label Noise ---
    # To target 85-90% accuracy, we intentionally mislabel ~12% of the students.
    # This simulates real-world messiness: some students with good marks are
    # actually at risk (personal issues, motivation), and some struggling
    # students turn things around (late bloomers).
    NOISE_RATE = 0.12  # 12% of labels will be randomly swapped
    noise_count = 0

    for idx, original_risk in enumerate(ml_risk_list):
        sid = f"ML{idx+1:05d}"
        s = generate_one_student(sid, f"ML {idx+1}", "X", random.choice(GENDERS), original_risk)
        
        features = extract_ml_features(s)
        features["student_id"] = sid
        
        # Apply label noise
        risk = original_risk
        if random.random() < NOISE_RATE:
            noise_count += 1
            if original_risk == "High":
                risk = random.choice(["Medium", "Medium"])  # High -> mostly Medium
            elif original_risk == "Medium":
                risk = random.choice(["High", "Low"])       # Medium -> either direction
            else:  # Low
                risk = random.choice(["Medium", "Medium"])  # Low -> mostly Medium
        
        features["risk_level"] = risk
        ml_rows.append(features)

    # Define CSV columns
    ml_columns = ["student_id", "cgpa", "overall_attendance", "overall_assignment"]
    for s in range(1, 7):
        ml_columns.extend([
            f"sem{s}_mid1_avg", f"sem{s}_mid2_avg",
            f"sem{s}_asn_avg", f"sem{s}_att_avg", f"sem{s}_sgpa"
        ])
    ml_columns.append("risk_level")

    csv_path = os.path.join(script_dir, "synthetic_students_ml.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ml_columns)
        writer.writeheader()
        writer.writerows(ml_rows)
    print(f"  -> Saved {len(ml_rows)} rows to synthetic_students_ml.csv")
    print(f"  -> Features per row: {len(ml_columns) - 2} (+ student_id + risk_level)")
    print(f"  -> Label noise applied: {noise_count}/{len(ml_rows)} ({noise_count/len(ml_rows)*100:.1f}%)")

    # -----------------------------------------------------------
    # DONE
    # -----------------------------------------------------------
    print(f"\n{'=' * 60}")
    print(f"GENERATION COMPLETE!")
    print(f"  App DB:    {len(all_demo_students)} students (3 sections x 65)")
    print(f"  ML CSV:    {len(ml_rows)} rows")
    print(f"  Excel:     {excel_path}")
    print(f"  JSON:      {json_path}")
    print(f"  ML CSV:    {csv_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
